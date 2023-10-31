import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class OutlierDetector:
    def __init__(self, excel_file_path=None, relative_size_threshold=5):
        """
        Initialize the OutlierDetector.

        Parameters:
        - excel_file_path (str, optional): Path to the input Excel file.
        - relative_size_threshold (float, optional): Threshold for determining relative size anomalies.
        """
        self.excel_file_path = excel_file_path
        self.relative_size_threshold = relative_size_threshold
        self.anomalies_dict = {}
        if excel_file_path:
            self.workbook = load_workbook(excel_file_path)
            self.sheet_names = self.workbook.sheetnames
            self.combined_data = self._combine_data_horizontally()
    
    def _combine_data_horizontally(self):
        """
        Combine data from all sheets in the workbook horizontally.

        Returns:
        - DataFrame: Combined data from all sheets.
        """
        combined_data = pd.DataFrame([cell for cell in col] for col in self.workbook[self.sheet_names[0]].iter_cols(values_only=True)).T
        for sheet in self.sheet_names[1:]:
            temp_df = pd.DataFrame([cell for cell in col] for col in self.workbook[sheet].iter_cols(values_only=True)).T
            temp_df = temp_df.drop(temp_df.columns[0], axis=1)
            combined_data = pd.concat([combined_data, temp_df], axis=1)
        return combined_data

    @staticmethod
    def convert_to_float(val):
        """
        Convert a value to float. If the value is a percentage (contains '%'), 
        it's converted to its decimal representation.

        Parameters:
        - val (str/int/float): Value to be converted.

        Returns:
        - float: Converted value or NaN if conversion is not possible.
        """
        try:
            if "%" in str(val):
                return float(val.rstrip('%')) / 100.0
            else:
                return float(val)
        except (ValueError, TypeError):
            return np.nan
    
    def detect_relative_median_anomalies(self):
        """
        Detect anomalies based on relative median for combined data.
        Anomalies are saved in the 'anomalies_dict' attribute.
        """
        for index, row in self.combined_data.iterrows():
            metric_name = row[0]
            if index == 0 or pd.isnull(metric_name):
                continue
            values = row[1:].apply(self.convert_to_float).dropna()
            rolling_median = values.rolling(window=3, center=True).median()
            relative_ratios = values / rolling_median
            anomalies = values[(relative_ratios > self.relative_size_threshold) | 
                               (relative_ratios < 1/self.relative_size_threshold)]
            if not anomalies.empty:
                self.anomalies_dict[metric_name] = anomalies
                
    def highlight_anomalies_in_excel(self, output_file_path=None):
        """
        Highlight detected anomalies in the Excel file with a red fill.

        Parameters:
        - output_file_path (str, optional): Path for saving the modified Excel file.
        """
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for sheet_name in self.sheet_names:
            worksheet = self.workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=worksheet.max_column):
                for cell in row:
                    metric = worksheet.cell(row=cell.row, column=1).value
                    cell_value = self.convert_to_float(cell.value)
                    if metric in self.anomalies_dict and cell_value in self.anomalies_dict[metric].values:
                        cell.fill = red_fill
        
        if output_file_path is None:
            output_file_path = self.excel_file_path.replace(".xlsx", "_modified.xlsx")
        self.workbook.save(output_file_path)

    def detect_anomalies_for_multiple_docs(self, directory_path):
        """
        Detect and highlight anomalies for multiple Excel documents in a specified directory.

        Parameters:
        - directory_path (str): Path to the directory containing the Excel files.
        """
        excel_files = [f for f in os.listdir(directory_path) if f.endswith('_cleaned.xlsx')]
        
        for excel_file in excel_files:
            self.excel_file_path = os.path.join(directory_path, excel_file)
            self.workbook = load_workbook(self.excel_file_path)
            self.sheet_names = self.workbook.sheetnames
            self.combined_data = self._combine_data_horizontally()
            self.detect_relative_median_anomalies()
            self.highlight_anomalies_in_excel()
