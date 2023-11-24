import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class OutlierDetector:
    def __init__(self, excel_file_path=None, threshold=0.5):
        self.excel_file_path = excel_file_path
        self.threshold = threshold
        self.data = None
        if excel_file_path:
            self._load_workbook()

    def _load_workbook(self):
        self.data = pd.read_excel(self.excel_file_path)

    def detect_relative_median_anomalies(self):
        anomalies = []
        
        # Convert non-numeric values to NaN for median calculation
        numeric_data = self.data.apply(pd.to_numeric, errors='coerce')
        
        # Compute the median for each column
        medians = numeric_data.median(axis=0, skipna=True)
        
        for col in self.data.columns:
            for index, value in self.data[col].items():
                # Check if the value is numeric and the median is not NaN
                if pd.notnull(value) and pd.notnull(medians[col]) and isinstance(value, (int, float)):
                    if abs(value - medians[col]) > self.threshold * medians[col]:
                        anomalies.append((index, self.data.columns.get_loc(col)))
        return anomalies


    def highlight_anomalies_in_excel(self, anomalies):
        workbook = load_workbook(self.excel_file_path)
        sheet = workbook.active
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        for row, col in anomalies:
            cell = sheet.cell(row=row+1, column=col+1)
            cell.fill = red_fill
        workbook.save(self.excel_file_path)

    def detect_anomalies_for_multiple_docs(self, directory_path):
        for filename in os.listdir(directory_path):
            if filename.endswith(".xlsx"):
                self.excel_file_path = os.path.join(directory_path, filename)
                self._load_workbook()
                anomalies = self.detect_relative_median_anomalies()
                self.highlight_anomalies_in_excel(anomalies)
