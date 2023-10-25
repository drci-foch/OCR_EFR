import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class OutlierDetector:
    def __init__(self, excel_file_path=None, relative_size_threshold=5):
        self.excel_file_path = excel_file_path
        self.relative_size_threshold = relative_size_threshold
        if excel_file_path:
            self.sheet_names = pd.ExcelFile(excel_file_path).sheet_names
            self.combined_data = self._combine_data_horizontally()
        self.anomalies_dict = {}
    
    def _combine_data_horizontally(self):
        combined_data = pd.read_excel(self.excel_file_path, sheet_name=self.sheet_names[0])
        for sheet in self.sheet_names[1:]:
            temp_df = pd.read_excel(self.excel_file_path, sheet_name=sheet)
            temp_df = temp_df.drop(temp_df.columns[0], axis=1)
            combined_data = pd.concat([combined_data, temp_df], axis=1)
        return combined_data
    
    @staticmethod
    def convert_to_float(val):
        try:
            if "%" in str(val):
                return float(val.rstrip('%')) / 100.0
            else:
                return float(val)
        except (ValueError, TypeError):
            return np.nan
    
    def detect_relative_median_anomalies(self):
        """
        Detects anomalies in the data based on the relative size compared to a rolling median.
        
        The method iterates through each row (metric) in the combined data. For each metric, 
        it calculates a rolling median using a window of 3 (the value itself, the previous value, 
        and the next value). 
        The relative size of each value to its corresponding rolling median 
        is then computed. 
        If the relative size exceeds the predefined threshold or is below the 
        inverse of the threshold, the value is flagged as an anomaly.
        
        Detected anomalies are stored in the `anomalies_dict` attribute of the class with the metric 
        name as the key and the anomalous values as the associated values.
        """
        for index, row in self.combined_data.iterrows():
            metric_name = row[0]
            if index == 0 or pd.isnull(metric_name):
                continue
            values = row[1:].apply(self.convert_to_float).dropna()
            rolling_median = values.rolling(window=3, center=True).median()
            relative_ratios = values / rolling_median
            anomalies = values[(relative_ratios > self.relative_size_threshold) | 
                               (relative_ratios < 1/self.relative_size_threshold)]
            if not anomalies.empty:
                self.anomalies_dict[metric_name] = anomalies
                
    def highlight_anomalies_in_excel(self, output_file_path=None):
        workbook = load_workbook(self.excel_file_path)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for sheet_name in self.sheet_names:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=worksheet.max_column):
                for cell in row:
                    metric = worksheet.cell(row=cell.row, column=1).value
                    if metric in self.anomalies_dict and cell.value in self.anomalies_dict[metric].values:
                        cell.fill = red_fill
        
        if output_file_path is None:
            output_file_path = self.excel_file_path.replace(".xlsx", "_modified.xlsx")
        workbook.save(output_file_path)

    def detect_anomalies_for_multiple_docs(self, directory_path):
        excel_files = [f for f in os.listdir(directory_path) if f.endswith('_cleaned.xlsx')]
        
        for excel_file in excel_files:
            self.excel_file_path = os.path.join(directory_path, excel_file)
            self.sheet_names = pd.ExcelFile(self.excel_file_path).sheet_names
            self.combined_data = self._combine_data_horizontally()
            self.detect_relative_median_anomalies()
            self.highlight_anomalies_in_excel()
