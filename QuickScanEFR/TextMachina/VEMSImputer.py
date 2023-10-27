import os
from openpyxl import load_workbook
import pandas as pd


class VEMSCorrector:
    def __init__(self, excel_file_path=None):
        """
        Initialize the VEMSCorrector.

        Parameters:
        - excel_file_path (str, optional): Path to the input Excel file.
        """
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.sheet_names = []
        self.combined_data = None
        if excel_file_path:
            self._load_workbook()

    def _load_workbook(self):
        """
        Load workbook and initialize necessary attributes.
        """
        self.workbook = load_workbook(self.excel_file_path)
        self.sheet_names = self.workbook.sheetnames
        self.combined_data = self._combine_data_horizontally()

    def _combine_data_horizontally(self):
        """
        Combine data from all sheets in the workbook horizontally.

        Returns:
        - DataFrame: Combined data from all sheets.
        """
        combined_data = pd.DataFrame(
            [cell for cell in col] for col in self.workbook[self.sheet_names[0]].iter_cols(values_only=True)).T
        for sheet in self.sheet_names[1:]:
            temp_df = pd.DataFrame([cell for cell in col]
                                   for col in self.workbook[sheet].iter_cols(values_only=True)).T
            temp_df = temp_df.drop(temp_df.columns[0], axis=1)
            combined_data = pd.concat([combined_data, temp_df], axis=1)
        return combined_data

    def compute_missing_vems(self):
        """
        Compute missing VEMS values for the combined data.
        """
        # Extracting rows related to VEMS and VEMS% from the combined data
        vems_row = self.combined_data[self.combined_data[0] == 'VEMS']
        vems_percent_row = self.combined_data[self.combined_data[0] == 'VEMS%']

        # Extract VEMS, VEMS%, and date values
        dates = self.combined_data.columns[1:].to_list()
        vems_values = vems_row.iloc[0, 1:].values
        vems_percent_values = [float(str(val).replace('%', '')) / 100 if pd.notnull(
            val) else None for val in vems_percent_row.iloc[0, 1:].values]

        # Placeholder for extrapolated values
        extrapolated_vems = []

        # Loop through each date
        for idx, (vems, vems_percent) in enumerate(zip(vems_values, vems_percent_values)):
            if pd.isnull(vems) and vems_percent is not None:
                # If VEMS is missing but VEMS% is provided

                # Find closest non-null VEMS values on both sides
                left_idx = idx - 1
                while left_idx >= 0 and pd.isnull(vems_values[left_idx]):
                    left_idx -= 1

                right_idx = idx + 1
                while right_idx < len(vems_values) and pd.isnull(vems_values[right_idx]):
                    right_idx += 1

                # Calculate difference in months for the dates
                current_date = pd.to_datetime(dates[idx])
                if 0 <= left_idx < len(vems_values):
                    left_date = pd.to_datetime(dates[left_idx])
                    left_diff_months = (
                        current_date.year - left_date.year) * 12 + current_date.month - left_date.month
                else:
                    left_diff_months = None

                if 0 <= right_idx < len(vems_values):
                    right_date = pd.to_datetime(dates[right_idx])
                    right_diff_months = (
                        right_date.year - current_date.year) * 12 + right_date.month - current_date.month
                else:
                    right_diff_months = None

                # If either left or right VEMS value is found and its date is less than 12 months apart, extrapolate VEMS
                if (0 <= left_idx < len(vems_values) and (left_diff_months is None or left_diff_months < 12)) or (0 <= right_idx < len(vems_values) and (right_diff_months is None or right_diff_months < 12)):
                    if left_diff_months is not None and right_diff_months is not None:
                        vems_estimated = ((vems_values[left_idx] * (vems_percent / vems_percent_values[left_idx])) +
                                          (vems_values[right_idx] * (vems_percent / vems_percent_values[right_idx]))) / 2
                    elif left_diff_months is not None:
                        vems_estimated = vems_values[left_idx] * \
                            (vems_percent / vems_percent_values[left_idx])
                    else:
                        vems_estimated = vems_values[right_idx] * \
                            (vems_percent / vems_percent_values[right_idx])

                    extrapolated_vems.append((idx, round(vems_estimated)))
                else:
                    extrapolated_vems.append((idx, None))
            else:
                extrapolated_vems.append((idx, None))

        # Update VEMS row with extrapolated values
        for idx, value in extrapolated_vems:
            if value is not None:
                vems_values[idx] = value

        # Update the dataframe with the new VEMS values
        self.combined_data.iloc[vems_row.index[0], 1:] = vems_values

    def compute_missing_vems_percent(self):
        """
        Compute missing VEMS% values based on provided VEMS values.
        """
        # Extracting rows related to VEMS and VEMS% from the combined data
        vems_row = self.combined_data[self.combined_data[0] == 'VEMS']
        vems_percent_row = self.combined_data[self.combined_data[0] == 'VEMS%']

        # Extract VEMS, VEMS%, and date values
        dates = self.combined_data.columns[1:].to_list()
        vems_values = vems_row.iloc[0, 1:].values
        vems_percent_values = [float(str(val).replace('%', '')) / 100 if pd.notnull(
            val) else None for val in vems_percent_row.iloc[0, 1:].values]

        # Placeholder for extrapolated VEMS% values
        extrapolated_vems_percent = []

        # Loop through each date
        for idx, (vems, vems_percent) in enumerate(zip(vems_values, vems_percent_values)):
            if pd.notnull(vems) and vems_percent is None:
                # If VEMS% is missing but VEMS is provided

                # Find closest non-null VEMS% values on both sides
                left_idx = idx - 1
                while left_idx >= 0 and vems_percent_values[left_idx] is None:
                    left_idx -= 1

                right_idx = idx + 1
                while right_idx < len(vems_percent_values) and vems_percent_values[right_idx] is None:
                    right_idx += 1

                # Calculate difference in months for the dates
                current_date = pd.to_datetime(dates[idx])
                if 0 <= left_idx < len(vems_percent_values):
                    left_date = pd.to_datetime(dates[left_idx])
                    left_diff_months = (
                        current_date.year - left_date.year) * 12 + current_date.month - left_date.month
                else:
                    left_diff_months = None

                if 0 <= right_idx < len(vems_percent_values):
                    right_date = pd.to_datetime(dates[right_idx])
                    right_diff_months = (
                        right_date.year - current_date.year) * 12 + right_date.month - current_date.month
                else:
                    right_diff_months = None

                # If either left or right VEMS% value is found and its date is less than 12 months apart, extrapolate VEMS%
                if (0 <= left_idx < len(vems_percent_values) and (left_diff_months is None or left_diff_months < 12)) or (0 <= right_idx < len(vems_percent_values) and (right_diff_months is None or right_diff_months < 12)):
                    if left_diff_months is not None and right_diff_months is not None:
                        vems_percent_estimated = ((vems / vems_values[left_idx] * vems_percent_values[left_idx]) +
                                                  (vems / vems_values[right_idx] * vems_percent_values[right_idx])) / 2
                    elif left_diff_months is not None:
                        vems_percent_estimated = vems / \
                            vems_values[left_idx] * \
                            vems_percent_values[left_idx]
                    else:
                        vems_percent_estimated = vems / \
                            vems_values[right_idx] * \
                            vems_percent_values[right_idx]

                    extrapolated_vems_percent.append(
                        (idx, round(vems_percent_estimated * 100, 2)))  # Convert back to percentage
                else:
                    extrapolated_vems_percent.append((idx, None))
            else:
                extrapolated_vems_percent.append((idx, None))

        # Update VEMS% row with extrapolated values
        for idx, value in extrapolated_vems_percent:
            if value is not None:
                vems_percent_values[idx] = f"{value}%"

        # Update the dataframe with the new VEMS% values
        self.combined_data.iloc[vems_percent_row.index[0],
                                1:] = vems_percent_values

    def save_combined_data(self, output_file_path=None):
        """
        Save the combined data with computed VEMS values to an Excel file.

        Parameters:
        - output_file_path (str, optional): Path for saving the combined data.
        """
        if output_file_path is None:
            output_file_path = self.excel_file_path.replace(
                ".xlsx", "_corrected.xlsx")
        self.combined_data.to_excel(output_file_path, index=False)

    def correct_multiple_docs(self, directory_path):
        """
        Correct VEMS values for multiple Excel documents in a specified directory.

        Parameters:
        - directory_path (str): Path to the directory containing the Excel files.
        """
        excel_files = [f for f in os.listdir(
            directory_path) if f.endswith('_cleaned_modified.xlsx')]

        for excel_file in excel_files:
            self.excel_file_path = os.path.join(directory_path, excel_file)
            self._load_workbook()
            self.compute_missing_vems()
            self.compute_missing_vems_percent()
            self.save_combined_data()
