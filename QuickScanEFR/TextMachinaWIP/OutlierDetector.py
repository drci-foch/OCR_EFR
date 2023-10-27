import os
from openpyxl.styles import PatternFill
from excel_utils import load_excel_workbook, load_workbook

class OutlierDetectorRefactored:
    def __init__(self, file_path=None, threshold=None):
        self.file_path = file_path
        self.threshold = threshold if threshold else 2.5

    @staticmethod
    def convert_to_float(value):
        """Convert a value to float. If the value is a percentage, it's converted to its decimal representation."""
        try:
            return float(value.strip('%')) / 100 if '%' in str(value) else float(value)
        except ValueError:
            return None

    def detect_relative_median_anomalies(self, series):
        """Detect anomalies based on relative size compared to the median."""
        series_float = series.apply(self.convert_to_float)
        median = series_float.median()
        anomalies = (series_float - median).abs() > self.threshold
        return anomalies

    def highlight_anomalies_in_excel(self, data, file_path):
        """Highlight detected anomalies in the Excel file with a red fill."""
        anomalies = data.apply(self.detect_relative_median_anomalies)
        workbook = load_workbook(file_path)
        sheet = workbook.active

        for col_idx, col_name in enumerate(anomalies.columns, 1):
            for row_idx, is_anomaly in enumerate(anomalies[col_name], 2):
                if is_anomaly:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        workbook.save(file_path)

    def detect_anomalies_for_multiple_docs(self, directory_path):
        """Detect and highlight anomalies for multiple Excel documents in a specified directory."""
        for filename in os.listdir(directory_path):
            if filename.endswith("_modified_corrected.xlsx"):
                file_path = os.path.join(directory_path, filename)
                data = load_excel_workbook(file_path)
                self.highlight_anomalies_in_excel(data, file_path)
                print(f"Processed anomalies for {file_path}")
