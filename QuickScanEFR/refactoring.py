# Let's start refactoring the scripts

# 1. Refactoring main.py
main_code_refactored = """
import os
from DateFormatter import DateFormatter
from PercentageExtractor import PercentageExtractor
from OutlierDetector import OutlierDetector
from VEMSCorrector import VEMSCorrector
from UnitCorrector import UnitCorrector
from MetricTypo import TypoCorrector
from ExcelCleanup import ExcelCleanup
from PDFPlumber import PDFProcessor

class MainPipeline:
    def __init__(self, input_directory='./pdf', output_directory='./excel'):
        self.input_directory = input_directory
        self.output_directory = output_directory
        self.standardized_metrics = ['VEMS', 'CV', 'CPT', 'VR', 'DLCO', 'KCO', 'VA']

        # Initialize helper classes
        self.date_formatter = DateFormatter()
        self.percentage_extractor = PercentageExtractor()
        self.outlier_detector = OutlierDetector()
        self.vems_imputer = VEMSCorrector()
        self.unit_corrector = UnitCorrector()
        self.typo_corrector = TypoCorrector(self.standardized_metrics)
        self.excel_cleaner = ExcelCleanup()
        self.pdf_processor = PDFProcessor(directory_path=input_directory)

    def process_pdfs(self):
        self.pdf_processor.process_directory()

    def format_dates(self):
        self.date_formatter.format_dates_in_excel(directory_path=self.output_directory)

    def extract_percentages(self):
        self.percentage_extractor.extract_percentages_from_excel(directory_path=self.output_directory)

    def detect_anomalies(self):
        self.outlier_detector.detect_anomalies_for_multiple_docs(directory_path=self.output_directory)

    def compute_missing_vems(self):
        self.vems_imputer.correct_multiple_docs(directory_path=self.output_directory)

    def correct_units(self):
        self.unit_corrector.correct_multiple_docs(directory_path=self.output_directory)

    def correct_metric_typos(self):
        # To be implemented based on the TypoCorrector class
        pass

    def clean_excel_files(self):
        self.excel_cleaner.clean_multiple_docs(directory_path=self.output_directory)

    def run(self):
        self.process_pdfs()
        self.format_dates()
        self.extract_percentages()
        self.detect_anomalies()
        self.compute_missing_vems()
        self.correct_units()
        self.correct_metric_typos()
        self.clean_excel_files()

    def cleanup(self):
        # Remove unnecessary files (e.g., intermediate Excel files) after processing
        pass

if __name__ == "__main__":
    pipeline = MainPipeline()
    pipeline.run()
"""

# Saving the refactored main.py
with open("./TextMachinaRefactored/main_refactored.py", "w") as file:
    file.write(main_code_refactored)

# 2. Refactoring PercentageExtractor.py
percentage_extractor_code_refactored = """
import os
import pandas as pd

class PercentageExtractor:
    def extract_percentages_from_row(self, row):
        values = row.values[1:]
        cleaned_values = []
        percentages = []
        for val in values:
            if isinstance(val, str) and "%" in val:
                percentages.append(val.replace('%', ''))
                cleaned_values.append(None)
            else:
                cleaned_values.append(val)
                percentages.append(None)
        return cleaned_values, percentages

    def process_dataframe_for_percentages(self, df):
        rows_to_append = []
        for index, row in df.iterrows():
            cleaned_values, percentages = self.extract_percentages_from_row(row)
            if any(percentages):
                new_row = [row.values[0] + ' (%)'] + percentages
                rows_to_append.append(new_row)
                df.iloc[index, 1:] = cleaned_values
        for new_row in rows_to_append:
            df = df.append(pd.Series(new_row, index=df.columns), ignore_index=True)
        return df

    def extract_percentages_from_excel(self, directory_path):
        for filename in os.listdir(directory_path):
            if filename.endswith(".xlsx") and not filename.endswith("_modified.xlsx"):
                file_path = os.path.join(directory_path, filename)
                df = pd.read_excel(file_path)
                df = self.process_dataframe_for_percentages(df)
                df.to_excel(file_path, index=False)

"""

# Saving the refactored PercentageExtractor.py
with open("./TextMachinaRefactored/PercentageExtractor_refactored.py", "w") as file:
    file.write(percentage_extractor_code_refactored)

# Continuing with refactoring DateFormatter.py
date_formatter_code_refactored = """
import os
import pandas as pd
import re
from datetime import datetime, timedelta

class DateFormatter:
    french_month_to_english = {
        'janvier': 'January',
        'février': 'February',
        'mars': 'March',
        'avril': 'April',
        'mai': 'May',
        'juin': 'June',
        'juillet': 'July',
        'août': 'August',
        'septembre': 'September',
        'octobre': 'October',
        'novembre': 'November',
        'décembre': 'December'
    }

    def extract_date_from_string(self, string):
        try:
            date_match = re.search(r"(\\d{1,2}) (janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre) (\\d{4})", string, re.IGNORECASE)
            day, month, year = date_match.groups()
            month = self.french_month_to_english[month.lower()]
            return f"{day} {month} {year}"
        except:
            return string

    def extract_relative_date(self, string, reference_date):
        # Assuming reference_date is in the format "dd Month yyyy"
        ref_date = datetime.strptime(reference_date, "%d %B %Y")
        if "day" in string:
            days = int(re.search(r"(\\d+) day", string).group(1))
            new_date = ref_date + timedelta(days=days)
            return new_date.strftime("%d %B %Y")
        # Add similar logic for other relative terms like "month", "year", etc.
        return string

    def format_dates_in_excel(self, directory_path):
        for filename in os.listdir(directory_path):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(directory_path, filename)
                df = pd.read_excel(file_path)
                df.iloc[0] = df.iloc[0].apply(self.extract_date_from_string)
                df.to_excel(file_path, index=False)
"""

# Saving the refactored DateFormatter.py
with open("./TextMachinaRefactored/DateFormatter_refactored.py", "w") as file:
    file.write(date_formatter_code_refactored)

# 4. Refactoring PDFPlumber.py
pdf_plumber_code_refactored = """
import os
import camelot
import pandas as pd

class PDFProcessor:
    def __init__(self, directory_path='./pdf'):
        self.directory_path = directory_path

    def process_pdf(self, pdf_path, output_path):
        tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
        with pd.ExcelWriter(output_path) as writer:
            for i, table in enumerate(tables):
                table.df.to_excel(writer, sheet_name=f'Sheet{i+1}', index=False)

    def process_directory(self):
        for filename in os.listdir(self.directory_path):
            if filename.endswith(".pdf"):
                pdf_path = os.path.join(self.directory_path, filename)
                output_path = os.path.join(self.directory_path.replace('/pdf', '/excel'), filename.replace('.pdf', '.xlsx'))
                self.process_pdf(pdf_path, output_path)
"""

# Saving the refactored PDFPlumber.py
with open("./TextMachinaRefactored/PDFPlumber_refactored.py", "w") as file:
    file.write(pdf_plumber_code_refactored)

# 5. Refactoring OutlierDetector.py
outlier_detector_code_refactored = """
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class OutlierDetector:
    def __init__(self, excel_file_path=None, threshold=0.5):
        self.excel_file_path = excel_file_path
        self.threshold = threshold
        self.combined_data = None
        if excel_file_path:
            self._load_workbook()

    def _load_workbook(self):
        self.combined_data = pd.read_excel(self.excel_file_path)

    def detect_relative_median_anomalies(self):
        anomalies = []
        medians = self.combined_data.median(axis=1, skipna=True)
        for index, row in self.combined_data.iterrows():
            for col, value in enumerate(row):
                if pd.notnull(value):
                    if abs(value - medians[index]) > self.threshold * medians[index]:
                        anomalies.append((index, col))
        return anomalies

    def highlight_anomalies_in_excel(self, anomalies):
        workbook = load_workbook(self.excel_file_path)
        sheet = workbook.active
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        for row, col in anomalies:
            cell = sheet.cell(row=row+1, column=col+1)
            cell.fill = red_fill
        workbook.save(self.excel_file_path)

    def detect_anomalies_for_multiple_docs(self, directory_path):
        for filename in os.listdir(directory_path):
            if filename.endswith(".xlsx"):
                self.excel_file_path = os.path.join(directory_path, filename)
                self._load_workbook()
                anomalies = self.detect_relative_median_anomalies()
                self.highlight_anomalies_in_excel(anomalies)
"""

# Saving the refactored OutlierDetector.py
with open("./TextMachinaRefactored/OutlierDetector_refactored.py", "w") as file:
    file.write(outlier_detector_code_refactored)


# Let's start refactoring the scripts, beginning with MetricTypo.py

# Refactoring MetricTypo.py
# The core logic of this script seems efficient. 
# I will make minor changes to enhance code readability and maintainability.

refactored_metric_typo_code = """
from typing import List

class TypoCorrector:
    def __init__(self, standardized_strings: List[str]):
        self.standardized_strings = standardized_strings

    @staticmethod
    def levenshtein_distance(s1: str, s2: str) -> int:
        if len(s1) > len(s2):
            s1, s2 = s2, s1

        distances = range(len(s1) + 1)
        for index2, char2 in enumerate(s2):
            new_distances = [index2 + 1]
            for index1, char1 in enumerate(s1):
                if char1 == char2:
                    new_distances.append(distances[index1])
                else:
                    new_distances.append(1 + min((distances[index1], distances[index1 + 1], new_distances[-1])))
            distances = new_distances

        return distances[-1]
    
    def correct(self, string: str, relevant_metrics: List[str], max_distance: int = 3) -> str:
        best_match = None
        best_distance = float('inf')
        for standard in relevant_metrics:
            distance = self.levenshtein_distance(string, standard)
            if distance < best_distance:
                best_distance = distance
                best_match = standard
        return best_match if best_distance <= max_distance else string
"""

# Saving the refactored MetricTypo.py script
refactored_metric_typo_path = "./TextMachinaRefactored/refactored_MetricTypo.py"
with open(refactored_metric_typo_path, "w") as file:
    file.write(refactored_metric_typo_code)

# Now, let's refactor UnitCorrector.py

# Refactoring UnitCorrector.py
# I'll optimize the code for better clarity and maintainability.

refactored_unit_corrector_code = """
import os
import pandas as pd
import re 

class UnitCorrector:
    def __init__(self, file_path=None):
        self.file_path = file_path
        self.workbook = None
        self.combined_data = None

    def load_workbook(self):
        self.workbook = pd.read_excel(self.file_path, None)
        dataframes = [self.workbook[sheet] for sheet in self.workbook.keys()]
        self.combined_data = pd.concat(dataframes, axis=1)
        
    def correct_values(self):
        correcting_row = self.combined_data.loc[self.combined_data[0].isin(['CV', 'CVF', 'VR Plethysmo', 'CPT Plethysmo', 'VEMS'])]
        if not correcting_row.empty:
            cvf_values = correcting_row.iloc[0, 1:].values
            corrected_values = [self._correct_unit(value) for value in cvf_values]
            self.combined_data.iloc[correcting_row.index[0], 1:] = corrected_values

    def _correct_unit(self, value):
        if pd.notnull(value):
            str_val = str(value).strip().lower()
            xlx_pattern = re.match(r"(\\d)[\\s]?l[\\s]?(\\d{2})", str_val)
            if xlx_pattern:
                return float(xlx_pattern.group(1) + "." + xlx_pattern.group(2)) * 1000
            elif "." in str_val and len(str_val.split(".")[0]) == 1 and len(str_val.split(".")[1]) == 2:
                return float(str_val) * 1000
            else:
                numeric_part = ''.join(filter(str.isdigit, str_val))
                unit = ''.join(filter(str.isalpha, str_val))
                return float(numeric_part) * 1000 if unit in ["l", "liters"] else float(numeric_part)
        return value

    def save_corrected_data(self, save_path=None):
        save_path = save_path or self.file_path.replace(".xlsx", "_cvf.xlsx")
        self.combined_data.to_excel(save_path, index=False)

    @staticmethod
    def correct_multiple_docs(directory_path):
        for filename in os.listdir(directory_path):
            if filename.endswith("_corrected.xlsx"):
                file_path = os.path.join(directory_path, filename)
                corrector = UnitCorrector(file_path)
                corrector.load_workbook()
                corrector.correct_values()
                corrector.save_corrected_data()
"""

# Saving the refactored UnitCorrector.py script
refactored_unit_corrector_path = "./TextMachinaRefactored/refactored_UnitCorrector.py"
with open(refactored_unit_corrector_path, "w") as file:
    file.write(refactored_unit_corrector_code)

refactored_unit_corrector_path


# Next, we'll refactor ExcelCleanup.py

# Refactoring ExcelCleanup.py
# I'll simplify the logic and structure for improved clarity.

refactored_excel_cleanup_code = """
import os
import pandas as pd
import re 

class ExcelCleanup:
    def __init__(self, file_path=None):
        self.file_path = file_path
        self.combined_data = None

    def load_workbook(self):
        self.combined_data = pd.read_excel(self.file_path)

    @staticmethod
    def format_percentage(value, is_first_column=False):
        if is_first_column or not isinstance(value, str):
            return value
        if re.match(r"^\\d+(\\.\\d+)?%$", value):
            return float(value.replace('%', ''))
        elif isinstance(value, (float, int)) and 0 < value <= 1:
            return value * 100
        return value

    def clean_percentage_values(self):
        rows_to_process = self.combined_data[self.combined_data.iloc[:, 0].str.endswith('%', na=False)]
        for _, row in rows_to_process.iterrows():
            for idx, col in enumerate(self.combined_data.columns):
                self.combined_data.at[row.name, col] = ExcelCleanup.format_percentage(self.combined_data.at[row.name, col], is_first_column=(idx == 0))

    def save_cleaned_data(self, save_path=None):
        save_path = save_path or self.file_path.replace(".xlsx", "_final.xlsx")
        self.combined_data.to_excel(save_path, index=False)

    @staticmethod
    def clean_multiple_docs(directory_path):
        for filename in os.listdir(directory_path):
            if filename.endswith("_cvf.xlsx"):
                file_path = os.path.join(directory_path, filename)
                cleaner = ExcelCleanup(file_path)
                cleaner.load_workbook()
                cleaner.clean_percentage_values()
                cleaner.save_cleaned_data()
"""

# Saving the refactored ExcelCleanup.py script
refactored_excel_cleanup_path = "./TextMachinaRefactored/refactored_ExcelCleanup.py"
with open(refactored_excel_cleanup_path, "w") as file:
    file.write(refactored_excel_cleanup_code)

# Now, let's refactor VEMSImputer.py

# Refactoring VEMSImputer.py
# I'll streamline the code and remove redundant parts for improved clarity.

refactored_vems_imputer_code = """
import os
import pandas as pd

class VEMSCorrector:
    def __init__(self, file_path=None):
        self.file_path = file_path
        self.combined_data = None

    def load_workbook(self):
        self.combined_data = pd.read_excel(self.file_path)

    @staticmethod
    def is_percent_value_usable(value):
        return isinstance(value, (int, float)) and 0 <= value <= 100

    def compute_missing_vems(self):
        vems_row = self.combined_data[self.combined_data[0] == 'VEMS'].iloc[0]
        for idx, val in enumerate(vems_row[1:]):
            if not self.is_percent_value_usable(val):
                self.combined_data.iat[vems_row.name, idx + 1] = None

    def save_combined_data(self, save_path=None):
        save_path = save_path or self.file_path.replace(".xlsx", "_vems.xlsx")
        self.combined_data.to_excel(save_path, index=False)

    @staticmethod
    def correct_multiple_docs(directory_path):
        for filename in os.listdir(directory_path):
            if filename.endswith("_cvf.xlsx"):
                file_path = os.path.join(directory_path, filename)
                corrector = VEMSCorrector(file_path)
                corrector.load_workbook()
                corrector.compute_missing_vems()
                corrector.save_combined_data()
"""

# Saving the refactored VEMSImputer.py script
refactored_vems_imputer_path = "./TextMachinaRefactored/refactored_VEMSImputer.py"
with open(refactored_vems_imputer_path, "w") as file:
    file.write(refactored_vems_imputer_code)

